import openpyxl
import os

wb = openpyxl.Workbook()
print(wb)

print(wb.sheetnames)

sheet = wb['Sheet']
print(sheet)

print(sheet['A1'].value)
print(sheet['A1'].value is None)

sheet['A1'] = 42
sheet['A2'] = 'Hello'

wb.save('example1.xlsx')

sheet2 = wb.create_sheet()
print(wb.sheetnames)
print(sheet2.title)
sheet2.title = 'My New Sheet Name'
print(wb.sheetnames)
wb.save('example2.xlsx')

wb.create_sheet(index=0, title='My other sheet')
wb.save('example3.xlsx')