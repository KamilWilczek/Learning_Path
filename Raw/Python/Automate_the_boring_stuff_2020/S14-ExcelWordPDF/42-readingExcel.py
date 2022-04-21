import openpyxl
import os
os.chdir(r'D:\Kursy\Automate the boring stuff\temp')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
# sheet = workbook.get_sheet_by_name('Sheet1')
sheet = workbook['Sheet1']
print(type(sheet))
# print(workbook.get_sheet_names())
print(workbook.sheetnames)
# print(sheet['A1'])
cell = sheet['A1']
print(cell.value)
print(str(cell.value))
print(sheet['A1'].value)
print(sheet['B1'].value)
print(sheet['C1'].value)
print(type(sheet['A1'].value))
print(type(sheet['B1'].value))
print(type(sheet['C1'].value))

print(sheet.cell(row=1, column=2))
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)